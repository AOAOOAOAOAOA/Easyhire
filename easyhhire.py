import random
import requests
import os
from datetime import datetime
from openpyxl import load_workbook
import telebot
from telebot import types   
import deepl as dep
import time

bot = telebot.TeleBot("6230489836:AAHnBg41o6RZbyLG2JjWeYfOfXRS5SWO0c8")
trans = dep.Translator("725d6dad-e54a-39d2-f167-16a0eec32055:fx")


anket_list = []



fnlog = "Enter.xlsx"
wblog =load_workbook(fnlog)
wslog = wblog["data"]

fnvak = "Vaks.xlsx"
wbvak = load_workbook(fnvak)
wsvak = wbvak["data"]

fn = "Anket.xlsx"
wb =load_workbook(fn)
ws = wb["data"]

lvl = 0

class Entered_in:
    def __init__(self,
                 user,
                 chatid):
        self.user = user
        self.chatid = chatid

entered = []
countentered = 0

Admin = "Oaeoaeoaoe"


Cat_list = []
Cat_list.append("информационные технологии")
Cat_list.append("маркетинг")
Cat_list.append("бьюти индустрия (красоты)")
Cat_list.append("гостиничный бизнес")
Cat_list.append("строительный бизнес, архитектура и дизайн")
Cat_list.append("туризм")
Cat_list.append("продажи")
Cat_list.append("рестораны и бары")
Cat_list.append("бизнес администрирование")
Cat_list.append("домашний персонал")
Cat_list.append("другие")

City_vac_list = []
City_vac_list.append("Лимасол")
City_vac_list.append("Строволос")
City_vac_list.append("Никосия")
City_vac_list.append("Ларнака")
City_vac_list.append("Лакатамия")
City_vac_list.append("Пафос")
City_vac_list.append("Като-Полемидия")
City_vac_list.append("Агландзия")
City_vac_list.append("Арадипу")
City_vac_list.append("Энгоми")
City_vac_list.append("Лация")
City_vac_list.append("Паралимни")
City_vac_list.append("Меса-Йитония")
City_vac_list.append("Айос-Атанасиос")
City_vac_list.append("Ермасойя")
City_vac_list.append("Айос-Дометиос")
City_vac_list.append("Ипсонас")
City_vac_list.append("Дали")


def Admin_Center(user):
    if (user == Admin):
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=2)
        Add_adm = types.KeyboardButton("Добавить Анкету")
        See_adm = types.KeyboardButton("Увидеть все анкеты")
        Change_adm = types.KeyboardButton("Изменить анкету")
        Writ_adm = types.KeyboardButton("Загрузить анкеты из даты")
        See_adm_vac = types.KeyboardButton("Увидеть все вакансии")
        Del_adm_vac = types.KeyboardButton("Удалить вакансию")
        Write_adm_vac = types.KeyboardButton("Загрузить вакансии из даты")
        markup.add(Add_adm, See_adm, Change_adm, Writ_adm, See_adm_vac, Del_adm_vac, Write_adm_vac)
        return markup



def classic(message):
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=2)
    search = types.KeyboardButton(str(trans.translate_text(text="[👁]Найти работника", target_lang=Find_Language(message))))
    do_anket = types.KeyboardButton(str(trans.translate_text(text="[🔧]Создать анкету", target_lang=Find_Language(message))))
    new_vak = types.KeyboardButton(str(trans.translate_text(text="[🔨]Создать вакансию", target_lang=Find_Language(message))))
    search_vak = types.KeyboardButton(str(trans.translate_text(text="[🫧]Найти вакансию", target_lang=Find_Language(message))))
    for vak in Vacantion_list:
        if (vak.user == message.from_user.username):
            new_vak = types.KeyboardButton(str(trans.translate_text(text="[🔨]Моя вакансия", target_lang=Find_Language(message))))
    for anket in anket_list:
        if (anket.user == message.from_user.username):
            do_anket = types.KeyboardButton(str(trans.translate_text(text="[🔧]Моя анкета", target_lang=Find_Language(message))))
    change = types.KeyboardButton(str(trans.translate_text(text="[🌐]Сменить язык", target_lang=Find_Language(message))))
    markup.add(search, do_anket, search_vak, new_vak, change)
    return markup







def AddVak(message):
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=2)
    back = types.KeyboardButton(str(trans.translate_text(text="[🏠]Вернуться в меню", target_lang=Find_Language(message))))
    new_vak = types.KeyboardButton(str(trans.translate_text(text="[🔨]Создать вакансию", target_lang=Find_Language(message))))
    #city = types.KeyboardButton("🔨Изменить город")
    #cat = types.KeyboardButton("🔨Изменить категорию")
    #change = types.KeyboardButton("🔨Пересобрать вакансию")
    delete = types.KeyboardButton(str(trans.translate_text(text="[🛑]Удалить вакансию", target_lang=Find_Language(message))))
    markup.add(back, new_vak, delete)
    return markup

def HowMuchVak(message):
    i = 0
    for vak in Vacantion_list:
        if (vak.user == message.from_user.username):
            i+=1
    return i

def FindIndexVak(User, idd):
    oud = 1
    o = 0
    for i in range(len(Vacantion_list)):
        print(f"{Vacantion_list[0].desc}")
        if (User == Vacantion_list[o].user):
            print(f"id {idd} = oud {oud}")
            if (idd == oud):
                print(Vacantion_list[o].desc)
                return i
            else:
                print("говно")
                oud += 1
        o += 1
    return o + 1
    print("нинадино")


def FindWhichVak(message, idd):
    idhere = 1
    idu = -1
    if (idd == "1"):
        idu = 1
    elif (idd == "2"):
        idu = 2
    elif (idd == "3"):
        idu = 3
    for vak in Vacantion_list:
        if (vak.user == message.from_user.username):
            if (idhere == idu):
                return vak
            else:
                idhere += 1
    return None




def ListOfCat():
    text = ""
    for cat in Cat_list:
        text += f"\n-" + cat 
    return text

def CheckListOfCat(message):
    for cat in Cat_list:
        if (message.text.upper() == cat.upper()):
            return True
    return False

def ListOfCityVAC():
    text = ""
    for city in City_vac_list:
        text += f"\n-" + city 
    return text

def CheckListOfCityVAC(message):
    for city in City_vac_list:
        if (message.text.upper() == city.upper()):
            return True
    return False





def DoChange(message):
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=2)
    namu = types.KeyboardButton(str(trans.translate_text(text="[🔧]Изменить Имя", target_lang=Find_Language(message))))
    dascu = types.KeyboardButton(str(trans.translate_text(text="[🔧]Изменить Описание", target_lang=Find_Language(message))))
    fotu = types.KeyboardButton(str(trans.translate_text(text="[🔧]Изменить Фото", target_lang=Find_Language(message))))
    video = types.KeyboardButton(str(trans.translate_text(text="[🔧]Изменить Видео", target_lang=Find_Language(message))))
    profu = types.KeyboardButton(str(trans.translate_text(text="[🔧]Изменить Професию", target_lang=Find_Language(message))))
    toun = types.KeyboardButton(str(trans.translate_text(text="[🔧]Изменить Город", target_lang=Find_Language(message))))
    back = types.KeyboardButton(str(trans.translate_text(text="[🏠]Вернуться в меню", target_lang=Find_Language(message))))
    anket_now = IsHisAnket(message.from_user.username)
    if (anket_now.active == True):
        act = types.KeyboardButton(str(trans.translate_text(text="[🛑]Деактивировать вашу анкету", target_lang=Find_Language(message))))
    else:
        act = types.KeyboardButton(str(trans.translate_text(text="[🛑]Активировать вашу анкету", target_lang=Find_Language(message))))
    
    markup.add(toun, namu, dascu, fotu, video ,profu,back, act )
    return markup

def Language(user):
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=2)
    Eng = types.KeyboardButton("🇬🇧 English")
    Ru = types.KeyboardButton("🇷🇺 руcкий")
    Ua = types.KeyboardButton("🇺🇦 Український")
    De = types.KeyboardButton("🇩🇪 Deutsch")
    El = types.KeyboardButton("🇬🇷 Ελληνική")
    Fr = types.KeyboardButton("🇫🇷 Français")
    markup.add(Eng, Ru, Ua, De, El, Fr)
    return markup


def eat(message):
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=2)
    find = types.KeyboardButton(str(trans.translate_text(text="[🔍]Перейти на его телеграм акаунт соискателя[🔎]", target_lang=Find_Language(message))))
    rew = types.KeyboardButton(str(trans.translate_text(text="[🏹]Продолжить поиск", target_lang=Find_Language(message))))
    back = types.KeyboardButton(str(trans.translate_text(text="[🏠]Вернуться в меню", target_lang=Find_Language(message))))
    markup.add(back, rew, find)
    return markup

def eatvak(message):
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=2)
    rew = types.KeyboardButton(str(trans.translate_text(text="[🏹]Продолжить поиск вакансий", target_lang=Find_Language(message))))
    back = types.KeyboardButton(str(trans.translate_text(text="[🏠]Вернуться в меню", target_lang=Find_Language(message))))
    markup.add(back, rew)
    return markup

class Anket:
    def __init__(self,
                 name,
                 description,
                 photo,
                 video,
                 profession,
                 city,
                 user):
        self.name = name
        self.description = description
        self.photo = photo
        self.video = video
        self.profession = profession
        self.city = city
        self.user = user
        self.active = True
        self.revs = []

class Founder:
    def __init__(self,
                user,
                city,
                prof):
        self.user = user
        self.city = city
        self.prof = prof

Founder_list = []

class Founder_vac:
    def __init__(self,
                 user,
                 city,
                 cat):
        self.user = user
        self.city = city
        self.cat = cat
    
Founder_vac_list = []



class CurrentUser:
    def __init__(self,
                 user, 
                 issearching):
        self.user = user
        self.issearching = issearching

Vacantion_list = []

class Vacantion:
    def __init__(self,
                 user,
                 cat,
                 city,
                 desc):
        self.user = user
        self.cat = cat
        self.city = city
        self.desc = desc


CurrentUser_list = []


Profession_list = []



#def ListOfProf():
#    was = []
#    text = ""
#    for profesion in Profession_list:
#        CanBe = True
#        for k in was:
#            if (profesion == k):
#                CanBe = False
#        if (CanBe == True):   
#            count = 0
#            for i in Profession_list:
#                if (profesion == i):
#                    count += 1
#            was.append(profesion)
#            text += f"\n{profesion} - {count} анкет с этой професией"
#    return text

class ListOnProfOne:
    def __init__(self,
                 prof,
                 count):
        self.prof = prof
        self.count = count


def ListOfProf():
    text = ""
    ARListOfProf = []
    for ank in anket_list:
        IsWas = False
        for i in range(len(ARListOfProf)):
            if (ARListOfProf[i].prof.upper() == ank.profession.upper()):
                IsWas = True
                ARListOfProf[i].count += 1
        if (IsWas == False):
            ARListOfProf.append(ListOnProfOne(prof=ank.profession, count= 1))
    for prof in ARListOfProf:
        text += f"\n{prof.prof} - {prof.count} анкет с этой професией"
    return text

class ListOnCityOne:
    def __init__(self,
                 city,
                 count):
        self.city = city
        self.count = count

City_list = []

def ListOfCity():
    text = ""
    ARListOfCity = []
    for ank in anket_list:
        IsWas = False
        for i in range(len(ARListOfCity)):
            if (ARListOfCity[i].city.upper() == ank.city.upper()):
                IsWas = True
                ARListOfCity[i].count += 1
        if (IsWas == False):
            ARListOfCity.append(ListOnCityOne(city=ank.city, count= 1))
    for city in ARListOfCity:
        text += f"\n{city.city} - {city.count} анкет с этого города"
    return text

def IsHisAnket(user):
    for anket in anket_list:
        if (anket.user == user):
            return anket

def SearchWithSameSett(user):
    for founder in Founder_list:
        if (founder.user == user):
            return founder
        
def SearchWithSameSettVak(user):
    for vak in Founder_vac_list:
        if (vak.user == user):
            return vak


def CurrentUserFound(user):
    for current_user in CurrentUser_list:
        if (current_user.user == user):
            return current_user

def CurrentUserDelete(user):
    i = 0
    for current_user in CurrentUser_list:
        if (current_user.user == user):
            return i
        else: 
            i += 1


def DataAdd(anket):
    index = str(anket_list.index(anket) + 2)
    ws["A" + index] = anket.user 
    ws["B" + index] = anket.name
    ws["C" + index] = anket.profession
    ws["D" + index] = anket.city
    ws["E" + index] = anket.description
    ws["F" + index] = anket.photo
    ws["G" + index] = anket.video
    ws["H" + index] = anket.active
    ws["J1"] = len(anket_list)
    
    wb.save(fn)

def DataAddVak(vak):
    index = str(Vacantion_list.index(vak) + 2)
    wsvak["A" + index] = vak.user
    wsvak["B" + index] = vak.cat 
    wsvak["C" + index] = vak.city
    wsvak["D" + index] = vak.desc
    wsvak["E" + index] = "#undeleted"
    wsvak["F2"] = len(Vacantion_list)

    print(wsvak["D" + index].value)

    wbvak.save(fnvak)    
    

def DataRead():
    indexx = 2
    count = ws["J1"].internal_value 
    for i in range(count):
            anket_list.append(Anket(name=ws["B" + str(indexx)].value, 
                                    description=ws["E" + str(indexx)].value, 
                                    photo=ws["F" + str(indexx)].value, 
                                    video=ws["G" + str(indexx)].value, 
                                    profession= ws["C" + str(indexx)].value, 
                                    city= ws["D" + str(indexx)].value, 
                                    user=ws["A" + str(indexx)].value))
            
            Profession_list.append(ws["C" + str(indexx)].value)
            City_list.append(ws["D" + str(indexx)].value)
            indexx += 1

def DataReadVak():
    indexx = 2
    count = wsvak["F2"].internal_value
    for i in range(count):
        if (wsvak["E" + str(indexx)].value != "#deleted"):
            Vacantion_list.append(Vacantion(user=wsvak["A" + str(indexx)].value,
                                            cat=wsvak["B" + str(indexx)].value,
                                            city=wsvak["C" + str(indexx)].value,
                                            desc=wsvak["D" + str(indexx)].value))
        else:
            pass
        indexx += 1
        
def Find_Language(message):
    
    if (message.from_user.username == Admin):
        return 'RU'
    else:
        indexx = wslog["F1"].value
        for i in range(indexx):
            if (f"@{message.from_user.username}" == wslog["A" + str(indexx)].value):
                return wslog["D" + str(indexx)].value
            
            
def Change_Language(message):
    
    if (message.from_user.username == Admin):
        return 'RU'
    else:
        indexx = wslog["F1"].value
        for i in range(indexx):
            if (f"@{message.from_user.username}" == wslog["A" + str(indexx)].value):
                lang = ""
                if (message.from_user.username != Admin):
                    if (message.text == "🇬🇧 English"): lang = "EN-GB"
                    if (message.text == "🇷🇺 руcкий"): lang = 'RU'
                    if (message.text == "🇺🇦 Український"): lang = "UK"
                    if (message.text == "🇩🇪 Deutsch"): lang = "DE"
                    if (message.text == "🇬🇷 Ελληνική"): lang = "EL"
                    if (message.text == "🇫🇷 Français"): lang = "FR"
                print(lang)
                wslog["D" + str(indexx)] = lang 
                wblog.save(fnlog)
                print (wslog["D" + str(indexx)].value)
        
       
def send_mes(message ,text, mark=0): bot.send_message(chat_id=message.chat.id, text=trans.translate_text(text=text, target_lang= Find_Language(message=message)))


def send_mes(message, text, mark): bot.send_message(chat_id=message.chat.id, text=trans.translate_text(text=text, target_lang= Find_Language(message=message)), reply_markup=mark)






@bot.message_handler(commands=["create_anket"])
def create_anket(message):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        #bot.send_message(chat_id=message.chat.id, text=f"Вам необходимо указать город на Кипре в котором вы б хотели-бы работать (нарпимер: Ланака, Лимассолл, Пафос, Никосия и т.д).\nЕсли вашего города тут нет, можете просто написать и он автоматически добавиться в список: \n" + ListOfCity())
        send_mes(message=message, text=f"Вам необходимо указать город на Кипре в котором вы б хотели-бы работать (нарпимер: Ланака, Лимассолл, Пафос, Никосия и т.д).\nЕсли вашего города тут нет, можете просто написать и он автоматически добавиться в список: \n" + ListOfCity(), mark=None )
        bot.register_next_step_handler(message, process_city)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        #send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")


def process_city(message):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        global city 
        city = message.text
        City_list.append(city)
        #bot.send_message(chat_id=message.chat.id, text=f"Пожайлуста, укажите работу на которую вы претендуете.\nЕсли вашей профессии тут нет, вы можете ее указать и она автоматически будет добавлена в список:\n" + ListOfProf())
        send_mes(message, text=f"Пожайлуста, укажите работу на которую вы претендуете.\nЕсли вашей профессии тут нет, вы можете ее указать и она автоматически будет добавлена в список:\n" + ListOfProf(), mark=None)
        bot.register_next_step_handler(message, process_profesion)
    except:
        bot.send_message(chat_id=message.chat.id, text="Неверно, попробуйте еще раз или чуть позже", reply_markup=classic(message))

def process_profesion(message):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        global profession
        profession = message.text
        Profession_list.append(profession)
        #bot.send_message(chat_id=message.chat.id, text="Укажите ваше имя и фамилию")
        send_mes(message, text= "Укажите ваше имя и фамилию", mark=0)
        bot.register_next_step_handler(message, process_name)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")

def process_name(message):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        global name
        name = message.text
        #bot.send_message(chat_id=message.chat.id, text="Для максимального полного наполнения вашей анкеты, рекомендуем внести следущие данные: \n-образование, \n-опыт работы, \n-возраст,\n-гражданство,\n-владение языками,\n-навычки,\n-ваши сильные стороны,\n-наличие прав и автомобиля,\n-статус прибываня на Кипре.\n\nТак-же вы можете добавить новые пункты по желанию")
        send_mes(message, mark=0, text="Для максимального полного наполнения вашей анкеты, рекомендуем внести следущие данные: \n-образование, \n-опыт работы, \n-возраст,\n-гражданство,\n-владение языками,\n-навычки,\n-ваши сильные стороны,\n-наличие прав и автомобиля,\n-статус прибываня на Кипре.\n\nТак-же вы можете добавить новые пункты по желанию" )
        bot.register_next_step_handler(message, process_description)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")

def process_description(message):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        global description
        description = message.text
        #bot.send_message(chat_id=message.chat.id, text="Дале загрузите Ваше фото, которое будут видеть работадатели")
        send_mes(message, mark=0, text="Дале загрузите Ваше фото, которое будут видеть работадатели")
        bot.register_next_step_handler(message, process_photo)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")

def process_photo(message):
    try:
        if isinstance(message.text, str):
            if (message.text.upper() == "СТОП"):
                send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
            else:
                send_mes(message, text="Не текст а фото, пожайлуста")
                bot.register_next_step_handler(message, process_photo)
        file_id = message.photo[-1].file_id
        file_info = bot.get_file(file_id)
        file = bot.download_file(file_info.file_path)
        with open(f"{file_id}.jpg", "wb") as f:
            f.write(file)
        #bot.send_message(chat_id=message.chat.id, text="Загрузите короткое видео в котором сообщите ваше имя а так-же расскажите кратко о вашем опыте работы для работодателя \n(если вы с компьютера, то файл должен быть в формате mp4)")
        send_mes(message, mark=0, text="Загрузите короткое видео длинной минутой, в котором сообщите ваше имя а так-же расскажите кратко о вашем опыте работы для работодателя \n(если вы с компьютера, то файл должен быть в формате mp4 до 20 mb)")
        bot.register_next_step_handler(message, process_video, file_id)
    except:
        #bot.send_message(chat_id=message.chat.id, text="Пожайлуста, отправте ФОТО")
        send_mes(message, mark=0, text="Пожайлуста, отправте ФОТО")
        bot.register_next_step_handler(message, process_photo)
    


def process_video(message, thing):
    try:
        if isinstance(message.text, str):
            if (message.text.upper() == "СТОП"):
                send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
            else:
                send_mes(message, text="Не текст а фото, пожайлуста")
                bot.register_next_step_handler(message, process_video, thing)
        video_id = message.video.file_id
        video = bot.get_file(video_id)
        video_url = f"https://api.telegram.org/file/bot{bot.token}/{video.file_path}"
        response = requests.get(video_url)
        if response.status_code == 200:
            with open(f"{video_id}.mp4", "wb") as f:
                f.write(response.content)
        anket = Anket(name, description, f"{thing}.jpg", f"{video_id}.mp4", profession, city, message.from_user.username)
        anket_list.append(anket)
        DataAdd(anket)
        send_mes(message, mark=classic(message), text=f"🎉Поздравляем! Ваша анкета готова и доступна для просмотра работадателем.\nДанную анкету вы можете редактировать, а так-же деактивировать в случае если это больше не актуально для вас" )
        #bot.send_message(chat_id=message.chat.id, reply_markup = classic(message.from_user.username), text=f"🎉Поздравляем! Ваша анкета готова и доступна для просмотра работадателем.\nДанную анкету вы можете редактировать, а так-же деактивировать в случае если это больше не актуально для вас")
    except:
        bot.send_message(chat_id=message.chat.id, text="Неправильный формат файла или длина видео дольше минуты, отправтье видео ")
        bot.register_next_step_handler(message, process_video, thing)

@bot.message_handler(commands=["get_random_anket"])
def get_random_anket(message):
    if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
    #bot.send_message(chat_id=message.chat.id, text="Укажите должность сотрудника, которого вы ищите исходя из следующего списка:\n" + ListOfProf())
    send_mes(message,mark=0, text= "Укажите должность сотрудника, которого вы ищите исходя из следующего списка:\n" + ListOfProf())
    bot.register_next_step_handler(message, get_city_anket)

def get_city_anket(message):
    if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
    #bot.send_message(chat_id=message.chat.id, text="Укажите город, который вы ищите из следующего списка:\n" + ListOfCity())
    send_mes(message, mark=0, text="Укажите город, который вы ищите из следующего списка:\n" + ListOfCity())
    bot.register_next_step_handler(message, find_random_anket, message.text, " ")

def find_random_anket(message, prof, cityy):
    #try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        if anket_list:
            BlackList = []
            Nothing = False
            while True:
                #мы проверили все анкеты?
                if(len(BlackList) == int(ws["J1"].value)):
                    Nothing = True
                    break
                
                index = random.randint(2, int(ws["J1"].value) + 1) #рандомный индекс
                
                was = False #был ли уже индекс в блэк лист?
                
                for i in BlackList:#проверка на блек лист
                    if (i == index):
                        was = True
                        
                if (was == False):#если не было
                    if(str(ws[f"H{index}"].value) == "True"): #если активированно
                        if(str(trans.translate_text(text=message.text, target_lang='RU')).upper() == str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper()): #если город подходит
                            gorone = f"{trans.translate_text(text=ws[f'C{index}'].value, target_lang='RU')}".upper()
                            gortwo = f"{trans.translate_text(text=prof, target_lang='RU')}".upper()
                            if (gorone == gortwo): #если професия подходит
                                CurrentUser_list.append(CurrentUser(message.from_user.username, ws[f"A{index}"])) #добавляем для кнопки "искать дальше"
                                #добовляем в групу фото и видео, и отпровляем
                                media_group = []
                                with open(ws[f"F{index}"].value, "rb") as f:
                                    media_group.append(telebot.types.InputMediaPhoto(f))
                                    with open(ws[f"G{index}"].value, "rb") as v:
                                        media_group.append(telebot.types.InputMediaVideo(v))
                                        bot.send_media_group(chat_id=message.chat.id, media = media_group)
                                #отпровляем инфу
                                send_mes(message, text=f'Имя: {ws[f"B{index}"].value}\n\nПрофесии: {ws[f"C{index}"].value} \nГород: {ws[f"D{index}"].value} \n\nОписание: \n{ws[f"E{index}"].value}', mark= eat(message))
                                

                                for pers in range(int(wslog["F1"].value)):
                                    print(f'@{message.from_user.username} == {wslog[f"A{pers + 1}"].value} // {wslog["F1"].value}')
                                    if (f"@{message.from_user.username}" == str(wslog[f"A{pers + 1}"].value)):
                                        print(f"{pers + 1}")
                                        wslog[f"G{pers + 1}"] = message.text
                                        wslog[f"H{pers + 1}"] = gorone
                                        print("fdgfdgs")
                                        wblog.save(fnlog)
                                
                                
                                
                                break
                            else: 
                                BlackList.append(index)

                        else: 
                            BlackList.append(index)
                            print(type(str(trans.translate_text(text=message.text, target_lang='RU')).upper()))
                            print(str(trans.translate_text(text=message.text, target_lang='RU')).upper())
                            print("=")
                            print(type(str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper()))
                            print(str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper())
                    else: 
                        BlackList.append(index)
                        print("актив")
            
            if (Nothing): #нет анкет
                send_mes(message,mark=0, text=f"Ввозможно вы ввели неправильно, пожайлуста, введите професию еще раз")
                #for i in BlackList:
                #    print(f'{ws[f"B{i}"].value} {ws[f"C{i}"].value.upper()} == {prof.upper()} /// {ws[f"D{i}"].value.upper()} == {message.text.upper()} /// {str(ws[f"H{index}"].value)} == true')
                bot.register_next_step_handler(message, get_city_anket)
               
            
            
            
            
            
            #IsNoAnket = False
            #BlackList = []
            #i = 0
            #while True:
            #    anket = random.choice(anket_list)
            #    print(f"{trans.translate_text(text=prof.upper(), target_lang=Find_Language(message))}  == {trans.translate_text(text=anket.profession.upper(), target_lang=Find_Language(message))}")
            #    print(f"{trans.translate_text(text=message.text.upper(), target_lang=Find_Language(message))} == {trans.translate_text(text=anket.city.upper(), target_lang=Find_Language(message))}")
            #    print("доходт")
            #    if (str(trans.translate_text(text=prof.upper(), target_lang="EN-GB"))== str(trans.translate_text(text=anket.profession.upper(), target_lang="EN-GB")) and str(trans.translate_text(text=message.text.upper(), target_lang="EN-GB")) == str(trans.translate_text(text=anket.city.upper(), target_lang="EN-GB")) and anket.active == True):
            #        print("DA")
            #        IsNoAnket = True
            #        break
            #    else:
            #        BlackList.append(anket)
            #    for ank in BlackList:
            #        if (ank == anket):
            #            i += 1
            #    if (i == len(anket_list)):
            #        break
            #    IsNoAnket
            #if (IsNoAnket == True):
            #   with open(anket.photo, "rb") as f:
            #        with open(anket.video, "rb") as v:
            #            
            #            media_group.append(telebot.types.InputMediaPhoto(f))
            #            media_group.append(telebot.types.InputMediaVideo(v))
            #            Founder_list.append(Founder(message.from_user.username, message.text, prof))
            #            CurrentUser_list.append(CurrentUser(message.from_user.username, anket.user))
            #            print (f"Добавлен юзер с ником {CurrentUserFound(message.from_user.username).issearching}")
            #            bot.send_media_group(chat_id=message.chat.id, media = media_group)
            #            send_mes(message, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description}", mark= eat(message))
            #            #bot.send_message(chat_id=message.chat.id, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description}", reply_markup= eat())
        #    else:
        #        #bot.send_message(chat_id=message.chat.id, text="Ввозможно вы ввели неправильно, попробуйте написать еще раз")
        #        #bot.send_message(chat_id=message.chat.id, text="Пожайлуста, введите професию")
        #        send_mes(message,mark=0, text=f"Ввозможно вы ввели неправильно, пожайлуста, введите професию еще раз")
        #        bot.register_next_step_handler(message, get_city_anket)
        #else:
        #    bot.send_message(chat_id=message.chat.id, text="No anket found", reply_markup= eat())
    #except:
        #bot.send_message(chat_id=message.chat.id, text="Неверно, попробуйте еще раз или чуть позже", reply_markup=classic(message))
        #get_random_anket(message)


@bot.message_handler(commands=["start"])
def start(message):
    bot.send_message(chat_id=message.chat.id, reply_markup=Language(message.from_user.username), text=f"🇬🇧 English?\n🇷🇺 руcкий?\n🇺🇦 Український?\n🇩🇪 Deutsch?\n🇬🇷 Ελληνική?\n🇫🇷 Français?" )
    print("")
    bot.register_next_step_handler(message, starte)

def starte (message):
    print("переход работает")
    curent_enter = Entered_in("@" + message.from_user.username, message.chat.id)
    entered.append(curent_enter)
    global countentered
    countentered += 1
    print(countentered)
    wslog["A" + str(countentered)] = curent_enter.user
    wslog["B" + str(countentered)] = curent_enter.chatid
    wslog["C" + str(countentered)] = datetime.now()
    lang = ""
    if (message.from_user.username != Admin):
        if (message.text == "🇬🇧 English"): lang = "EN-GB"
        if (message.text == "🇷🇺 руcкий"): lang = 'RU'
        if (message.text == "🇩🇪 Deutsch"): lang = "DE"
        if (message.text == "🇺🇦 Український"): lang = "UK"
        if (message.text == "🇬🇷 Ελληνική"): lang = "EL"
        if (message.text == "🇫🇷 Français"): lang = "FR"
    print (lang)
    print(message.text)
    wslog["D" + str(countentered)] = lang
    wslog["F1"] = countentered
    wblog.save(fnlog)
    print("усе загрузил")
    print(wslog["A" + str(countentered)].value)
    print(Find_Language(message))
    send_mes(message, text=f"Здраствуйте, это бесплатная бета-версия бота \"EasyHire\". \nДолжен отметить, что в конечной версии бот будет выглядеть более совершенным и в то же время станет платным. Все анкеты и вакансии будут сохранены в конечного бота.\nИнструкция:\n[👁]Найти работника - этой опцией вы можете воспользоваться, если вы ищите сотрудника\n[🔧]Создать анкету - вы можете использовать данный пункт для поиска работы\n[🔨]Создать вакансию - в этом разделе работадатели, имеют возможность, сообщить об открытии вакансии и требовании к соискателю \n[🫧]Найти вакансию - вы можете воспользоваться этим пунктом, чтобы найти необходимую вам работу", mark= classic(message))
    print("все сказал")
    

@bot.message_handler(commands=["create_vacantions"])
def create_vacantions(message):
    try:
        if (message.text.upper() == "СТОП"):
            #send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        if (HowMuchVak(message) < 3):
            print(HowMuchVak(message))
            print(message.from_user.username)
            current_vacantion = Vacantion(user=message.from_user.username, cat= "null", city= "null", desc= "null")
            #bot.send_message(chat_id=message.chat.id, text="Выберите из представленных категорий, необходимую:" + ListOfCat())
            send_mes(message,mark=0,text= "Выберите из представленных категорий, необходимую:" + ListOfCat())
            bot.register_next_step_handler(message, vacantions_procc_cat, current_vacantion)
        else:
            send_mes(message, mark=AddVak(message), text="Простите, но у вас уже есть 3 вакансии, это пока максимум, если требуеться больше, пишите на @EasyHireHELP")
           # bot.send_message(chat_id=message.chat.id, text="Простите, но у вас уже есть 3 вакансии, это пока максимум, если требуеться больше, пишите на @EasyHireHELP", reply_markup=AddVak())
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, create_vacantions)

def vacantions_procc_cat(message, current_vacantion):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        if (CheckListOfCat(message=message)):
            #bot.send_message(chat_id=message.chat.id, text=f"Вы выбрали категорию:{message.text}")
            send_mes(message, mark=0,text=f"Вы выбрали категорию:{message.text}")
            current_vacantion.cat = message.text
            #bot.send_message(chat_id=message.chat.id, text="Выберите город:" + ListOfCityVAC())
            send_mes(message,mark=0, text="Выберите город:" + ListOfCityVAC())
            bot.register_next_step_handler(message, vacantions_city_cat, current_vacantion)
        else:
            #bot.send_message(chat_id=message.chat.id, text=f"Неправильно набранная категория, попробуйте еще раз")
            send_mes(message,mark=0, text=f"Неправильно набранная категория, попробуйте еще раз")
            bot.register_next_step_handler(message, vacantions_procc_cat, current_vacantion)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, vacantions_procc_cat, current_vacantion)

def vacantions_city_cat(message, current_vacantion):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        if (CheckListOfCityVAC(message=message)):
            current_vacantion.city = message.text
            #bot.send_message(chat_id=message.chat.id, text=f"Вы выбрали город:{message.text}")
            send_mes(message,mark=0, text=f"Вы выбрали город:{message.text}")
            #bot.send_message(chat_id=message.chat.id, text="Опишите обязанности и требования к кандидату, а так же укажите ваши контактные данные \n(до 4000 символов)")
            send_mes(message,mark=0, text="Опишите обязанности и требования к кандидату, а так же укажите ваши контактные данные \n(до 4000 символов)")
            bot.register_next_step_handler(message, vacantions_desc_cat, current_vacantion)
        else:
            #bot.send_message(chat_id=message.chat.id, text="Неправильно набранный город, попробуйте еще раз")
            send_mes(message,mark=0, text="Неправильно набранный город, попробуйте еще раз")
            bot.register_next_step_handler(message, vacantions_city_cat, current_vacantion)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, vacantions_city_cat, current_vacantion)

def vacantions_desc_cat(message, current_vacantion):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        if (len(message.text) < 4000):
            current_vacantion.desc = f"Заказщик: @{current_vacantion.user} \nКатегория: {current_vacantion.cat} \nГород: {current_vacantion.city} \n\n{message.text}"
            Vacantion_list.append(current_vacantion)
            #bot.send_message(chat_id=message.chat.id, text=current_vacantion.desc)
            send_mes(message,mark=0,text=current_vacantion.desc )
            DataAddVak(current_vacantion)
            #bot.send_message(chat_id=message.chat.id, text=f"🎉Поздравляем! Вы создали вакансию.\n Будем вам благодарны если вы удалите вакансию, после того как найдете необходимого вам сотрудника (чтобы вас не беспокоили другие соискатели)", reply_markup= classic(message.from_user.username))
            send_mes(message, text=f"🎉Поздравляем! Вы создали вакансию.\n Будем вам благодарны если вы удалите вакансию, после того как найдете необходимого вам сотрудника (чтобы вас не беспокоили другие соискатели)", mark= classic(message))
        else:
            #bot.send_message(chat_id=message.chat.id, text="ДО 4000 СИМВОЛОВ")
            send_mes(message, mark=0,text="ДО 4000 СИМВОЛОВ")
            bot.register_next_step_handler(message, vacantions_desc_cat, current_vacantion)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, vacantions_desc_cat, current_vacantion)


@bot.message_handler(commands=["find_vacantion"])
def find_vacantion(message):
    try:
        current_founder = Founder_vac(user= message.from_user.username, city= "null", cat = "null")
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        #bot.send_message(chat_id=message.chat.id, text="Впишите представленый выбранный вами категорию ниже:"+ ListOfCat())
        send_mes(message, mark=0,text="Впишите представленый выбранный вами категорию ниже:"+ ListOfCat())
        bot.register_next_step_handler(message, find_vacantion_procc_prof, current_founder)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, find_vacantion)
    

def find_vacantion_procc_prof(message, current_founder):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        if (CheckListOfCat(message)):
            current_founder.cat = message.text
            #bot.send_message(chat_id=message.chat.id, text="Выберите город:"+ ListOfCityVAC())
            send_mes(message, mark=0,text="Выберите город:"+ ListOfCityVAC())
            bot.register_next_step_handler(message, find_vacantion_procc_city, current_founder)
        else:
           #bot.send_message(chat_id=message.chat.id, text="Неправильно набранная категория, попробуйте еще раз")
            send_mes(message, mark=0,text="Неправильно набранная категория, попробуйте еще раз")
            bot.register_next_step_handler(message, find_vacantion)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, find_vacantion_procc_prof)

def find_vacantion_procc_city(message, current_founder):
    try:
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        if (CheckListOfCityVAC(message)):
            current_founder.city = message.text
            i = 0
            IsNoVac = False
            BlackList = []
            while True:
                IsWasVac = False
                CountWas = 0
                vacantion = random.choice(Vacantion_list)
                print(f"{trans.translate_text(text=current_founder.cat.upper(), target_lang=Find_Language(message))} == {trans.translate_text(text=vacantion.cat.upper(), target_lang=Find_Language(message))}")
                print(f"{trans.translate_text(text=current_founder.city.upper(), target_lang=Find_Language(message))} ==  {trans.translate_text(text=vacantion.city.upper(), target_lang=Find_Language(message))}")
                if (str(trans.translate_text(text=current_founder.cat.upper(), target_lang="EN-GB")) == str(trans.translate_text(text=vacantion.cat.upper(), target_lang="EN-GB")) and str(trans.translate_text(text=current_founder.city.upper(), target_lang="EN-GB")) == str(trans.translate_text(text=vacantion.city.upper(), target_lang="EN-GB"))):
                    IsNoVac = True
                    break
                else:
                    for vac in BlackList:
                        if (vac == vacantion):
                            IsWasVac = True
                    if (IsWasVac == False):
                        BlackList.append(vacantion)
                if (len(BlackList) == len(Vacantion_list)):
                    IsNoVac = False
                    break
            if (IsNoVac):
                #bot.send_message(chat_id=message.chat.id, text=vacantion.desc, reply_markup= eatvak())
                send_mes(message, text=vacantion.desc, mark=eatvak(message))
                Founder_vac_list.append(current_founder)
            elif(IsNoVac == False):
                #bot.send_message(chat_id=message.chat.id, text="простите, по вашим запросам анкет еще нет", reply_markup=classic(message.from_user.username))
                send_mes(message, text="простите, по вашим запросам анкет еще нет", mark=classic(message))
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, find_vacantion_procc_city)
    

@bot.message_handler()
def everything(message):
    text = message.text

    if (text.upper() == "пошел нахуй".upper()):
        bot.send_message(chat_id=message.chat.id, reply_markup = None, text="Сам иди" )
    
    if (message.from_user.username == Admin):
        bot.send_message(chat_id=message.chat.id, reply_markup = Admin_Center(message.from_user.username), text="Да здластвует нас калоль" )
    if (text == str(trans.translate_text(text="[🔧]Создать анкету", target_lang=Find_Language(message)))):
        create_anket(message)
    if (text == str(trans.translate_text(text="[👁]Найти работника", target_lang=Find_Language(message)))):
        get_random_anket(message)
    if (text == str(trans.translate_text(text="[🔨]Создать вакансию", target_lang=Find_Language(message)))):
        create_vacantions(message)
    if(text == str(trans.translate_text(text="[🔨]Моя вакансия", target_lang=Find_Language(message)))):
        SeeAllVaks(message=message)
    if (text == str(trans.translate_text(text="[🫧]Найти вакансию", target_lang=Find_Language(message)))):
        find_vacantion(message)
    if (text == str(trans.translate_text(text="[🏹]Продолжить поиск вакансий", target_lang=Find_Language(message)))):
        if (message.text.upper() == "СТОП"):
            send_mes(message, text=f"Мы вернули вас в меню", mark=classic(message))
        if anket_list:
            BlackList = []
            Nothing = False
            while True:
                #мы проверили все анкеты?
                if(len(BlackList) == int(ws["J1"].value)):
                    Nothing = True
                    break
                
                index = random.randint(2, int(ws["J1"].value) + 1) #рандомный индекс
                
                was = False #был ли уже индекс в блэк лист?
                
                for i in BlackList:#проверка на блек лист
                    if (i == index):
                        was = True
                        
                if (was == False):#если не было
                    if(str(ws[f"H{index}"].value) == "True"): #если активированно
                        if(str(trans.translate_text(text=message.text, target_lang='RU')).upper() == str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper()): #если город подходит
                            gorone = f"{trans.translate_text(text=ws[f'C{index}'].value, target_lang='RU')}"
                            print(type(gorone))
                            gorone = gorone.upper()
                            gortwo = f"{trans.translate_text(text=prof, target_lang='RU')}".upper()
                            print(type(gortwo))
                            gortwo = gortwo.upper()
                            print("------")
                            if (gorone == gortwo): #если професия подходит
                                print("++++")
                                CurrentUser_list.append(CurrentUser(message.from_user.username, ws[f"A{index}"])) #добавляем для кнопки "искать дальше"
                                print("////")
                                #print (f'Добавлен юзер с ником {ws["A" + index].value}') #для дебага
                                print("((()))")
                                #добовляем в групу фото и видео, и отпровляем
                                media_group = []
                                with open(ws[f"F{index}"].value, "rb") as f:
                                    media_group.append(telebot.types.InputMediaPhoto(f))
                                    print("*****")
                                    with open(ws[f"G{index}"].value, "rb") as v:
                                        media_group.append(telebot.types.InputMediaVideo(v))
                                        print(":::::::::::")
                                        bot.send_media_group(chat_id=message.chat.id, media = media_group)
                                print("LLLLLL")
                                #отпровляем инфу
                                send_mes(message, text=f'Имя: {ws[f"B{index}"].value}\n\nПрофесии: {ws[f"C{index}"].value} \nГород: {ws[f"D{index}"].value} \n\nОписание: \n{ws[f"E{index}"].value}', mark= eat(message))
                                break
                            else: 
                                BlackList.append(index)
                                print("----------------------------")
                                print(type(gorone))
                                print(gorone)
                                print("=")
                                print(type(gortwo))
                                print(gortwo)
                                print("----------------------------")

                        else: 
                            BlackList.append(index)
                            print(type(str(trans.translate_text(text=message.text, target_lang='RU')).upper()))
                            print(str(trans.translate_text(text=message.text, target_lang='RU')).upper())
                            print("=")
                            print(type(str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper()))
                            print(str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper())
                    else: 
                        BlackList.append(index)
                        print("актив")
            
            if (Nothing): #нет анкет
                send_mes(message,mark=0, text=f"Ввозможно вы ввели неправильно, пожайлуста, введите професию еще раз")
                #for i in BlackList:
                #    print(f'{ws[f"B{i}"].value} {ws[f"C{i}"].value.upper()} == {prof.upper()} /// {ws[f"D{i}"].value.upper()} == {message.text.upper()} /// {str(ws[f"H{index}"].value)} == true')
                bot.register_next_step_handler(message, get_city_anket)
                

    if (text == str(trans.translate_text(text="[🛑]Удалить вакансию", target_lang=Find_Language(message)))):
        remake_vac(message)
    if (text == str(trans.translate_text(text="[🔍]Перейти на его телеграм акаунт соискателя[🔎]", target_lang=Find_Language(message)))):
        curr = CurrentUserFound(message.from_user.username)
        print (f"найден юзеру {CurrentUserFound(message.from_user.username).user} этот {curr.issearching}")
        #bot.send_message(chat_id=message.chat.id, reply_markup = eat(), text=f"Телеграм: @{curr.issearching}" )
        send_mes(message, mark=eat(message), text=f"Телеграм: @{curr.issearching}")
        print (f"удален юзеру {CurrentUserFound(message.from_user.username).user} этот {curr.issearching}")
        CurrentUser_list.pop(CurrentUserDelete(message.from_user.username))
        print (f"теперь точно")
    if (text == str(trans.translate_text(text="[🏠]Вернуться в меню", target_lang=Find_Language(message)))):
        #bot.send_message(chat_id=message.chat.id, reply_markup = classic(message.from_user.username), text=f"Меню")
        send_mes(message, text=f"Меню", mark=classic(message))
    if (text == str(trans.translate_text(text="[🌐]Сменить язык", target_lang=Find_Language(message)))):
        bot.send_message(chat_id=message.chat.id, reply_markup=Language(message), text=f"🇬🇧 English?\n🇷🇺 руcкий?\n🇺🇦 Український?\n🇩🇪 Deutsch?\n🇬🇷 Ελληνική?\n🇫🇷 Français?" )
        bot.register_next_step_handler(message, Change_langue)
        #Change_langue(message)
    if (text == str(trans.translate_text(text="[🔧]Моя анкета", target_lang=Find_Language(message)))):
        for anket in anket_list:
            if (message.from_user.username == anket.user):
                if (anket.active == True):
                    with open(anket.photo, "rb") as f:
                        with open(anket.video, "rb") as v:
                            media_group = []
                            media_group.append(telebot.types.InputMediaPhoto(f))
                            media_group.append(telebot.types.InputMediaVideo(v))
                            current_user = anket.user
                            bot.send_media_group(chat_id=message.chat.id, media = media_group)
                            #bot.send_message(chat_id=message.chat.id, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Активированая", reply_markup= DoChange(user=message.from_user.username))
                            send_mes(message, mark=DoChange(message), text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Активированая")
                if (anket.active == False):
                    with open(anket.photo, "rb") as f:
                        with open(anket.video, "rb") as v:
                            media_group = []
                            media_group.append(telebot.types.InputMediaPhoto(f))
                            media_group.append(telebot.types.InputMediaVideo(v))
                            current_user = anket.user
                            bot.send_media_group(chat_id=message.chat.id, media = media_group)
                            send_mes(message, mark=DoChange(message), text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Деактивированая")
                            #bot.send_message(chat_id=message.chat.id, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Деактивированая", reply_markup= DoChange(user=message.from_user.username))
    
    if (text == str(trans.translate_text(text="[🔧]Изменить Имя", target_lang=Find_Language(message)))):
        #bot.send_message(chat_id=message.chat.id, text="Отправтье ваше новое имя")
        send_mes(message, mark=0, text="Отправтье ваше новое имя")
        bot.register_next_step_handler(message, change_procc_name)

    if (text == str(trans.translate_text(text="[🔧]Изменить Описание", target_lang=Find_Language(message)))):
        #bot.send_message(chat_id=message.chat.id, text="Добавтье Ваши данные:\n-образование \n-опыт работы \n-возраст,\n-гражданство,\n-владение языками,\n-навычки,\n-ваши сильные стороны,\n-наличие прав и автомобиля,\n-статус прибываня на кипре")
        send_mes(message, mark=0,text="Добавтье Ваши данные:\n-образование \n-опыт работы \n-возраст,\n-гражданство,\n-владение языками,\n-навычки,\n-ваши сильные стороны,\n-наличие прав и автомобиля,\n-статус прибываня")
        #bot.send_message(chat_id=message.chat.id, text="Отправтье ваше новое описание:")
        send_mes(message, mark=0, text="Отправтье ваше новое описание:")
        bot.register_next_step_handler(message, change_procc_desc)
    
    if (text == str(trans.translate_text(text="[🔧]Изменить Фото", target_lang=Find_Language(message)))):
        #bot.send_message(chat_id=message.chat.id, text="Отправтье ваше фото")
        send_mes(message, mark=0, text="Отправтье ваше фото")
        bot.register_next_step_handler(message, change_procc_photo)
    
    if (text == str(trans.translate_text(text="[🔧]Изменить Професию", target_lang=Find_Language(message)))):
        #bot.send_message(chat_id=message.chat.id, text="Отправтье вашу новою професию")
        send_mes(message, mark=0, text="Отправтье вашу новою професию")
        bot.register_next_step_handler(message, change_procc_prof)
    
    if (text == str(trans.translate_text(text="[🔧]Изменить Видео", target_lang=Find_Language(message)))):
        #bot.send_message(chat_id=message.chat.id, text="Отправтье ваше новое видео")
        send_mes(message, mark=0, text="Отправтье ваше новое видео")
        bot.register_next_step_handler(message, change_procc_video)
    
    if (text == str(trans.translate_text(text="[🔧]Изменить Город", target_lang=Find_Language(message)))):
        #bot.send_message(chat_id=message.chat.id, text="Отправтье ваш новый город")
        send_mes(message, mark=0, text="Отправтье ваш новый город")
        bot.register_next_step_handler(message, change_procc_city)

    if (text == str(trans.translate_text(text="[🛑]Деактивировать вашу анкету", target_lang=Find_Language(message)))):
        anket_now = IsHisAnket(message.from_user.username)
        anket_now.active = False
        ws["H" + str(anket_list.index(anket_now) + 2)] = anket_now.active
        wb.save(fn)
        #bot.send_message(chat_id=message.chat.id, text="Поздравляем! Вы поменяли Деактивировали вашу анкету", reply_markup =DoChange(message.from_user.username))
        send_mes(message, text="Поздравляем! Вы поменяли Деактивировали вашу анкету", mark=DoChange(message))
    
    if (text == str(trans.translate_text(text="[🛑]Активировать вашу анкету", target_lang=Find_Language(message)))):
        anket_now = IsHisAnket(message.from_user.username)
        anket_now.active = True
        ws["H" + str(anket_list.index(anket_now) + 2)] = anket_now.active
        wb.save(fn)
        send_mes(message, text="Поздравляем! Вы поменяли Активировали вашу анкету", mark=DoChange(message))
        #bot.send_message(chat_id=message.chat.id, text="Поздравляем! Вы поменяли Активировали вашу анкету", reply_markup =DoChange(message.from_user.username))


    if (text == str(trans.translate_text(text="[🏹]Продолжить поиск", target_lang=Find_Language(message)))):
        
        pprof = ""
        ccity = ""
        print("раз")
        for i in range(int(wslog["F1"].value)):
            #print(f'{f"{wslog[f'A{i + 1}'].value}"} == {f"@{message.from_user.username}"}')
            if (f"{wslog[f'A{i + 1}'].value}" == f"@{message.from_user.username}"):
                print ("тест")
                print(str(wslog[f"G{i + 1}"].value))
                print(i + 1)
                ccity = str(wslog[f"G{i + 1}"].value)
                pprof = str(wslog[f"H{i + 1}"].value)
        print("два")
        if anket_list:
            BlackList = []
            Nothing = False
            while True:
                print("три")
                #мы проверили все анкеты?
                if(len(BlackList) == int(ws["J1"].value)):
                    Nothing = True
                    break
                
                index = random.randint(2, int(ws["J1"].value) + 1) #рандомный индекс
                
                was = False #был ли уже индекс в блэк лист?
                
                for i in BlackList:#проверка на блек лист
                    if (i == index):
                        was = True
                print("четыре")

                if (was == False):#если не было
                    print("одинин")
                    if(str(ws[f"H{index}"].value) == "True"): #если активированно
                        print("дванин")
                        print(ccity)
                        print()
                        print(ws[f"D{index}"].value)
                        if(str(trans.translate_text(text=ccity, target_lang='RU')).upper() == str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper()): #если город подходит
                            print("тринин")
                            gorone = f"{trans.translate_text(text=ws[f'C{index}'].value, target_lang='RU')}".upper()
                            gortwo = f"{trans.translate_text(text=pprof, target_lang='RU')}".upper()
                            print("виви")
                            if (gorone == gortwo): #если професия подходит
                                print("пять")
                                CurrentUser_list.append(CurrentUser(message.from_user.username, ws[f"A{index}"])) #добавляем для кнопки "искать дальше"
                                #добовляем в групу фото и видео, и отпровляем
                                media_group = []
                                with open(ws[f"F{index}"].value, "rb") as f:
                                    media_group.append(telebot.types.InputMediaPhoto(f))
                                    with open(ws[f"G{index}"].value, "rb") as v:
                                        media_group.append(telebot.types.InputMediaVideo(v))
                                        bot.send_media_group(chat_id=message.chat.id, media = media_group)
                                #отпровляем инфу
                                send_mes(message, text=f'Имя: {ws[f"B{index}"].value}\n\nПрофесии: {ws[f"C{index}"].value} \nГород: {ws[f"D{index}"].value} \n\nОписание: \n{ws[f"E{index}"].value}', mark= eat(message))
                                indind = 0

                                for pers in range(int(wslog["F1"])):
                                    if (f"@{message.from_user.username}" == str(wslog[f"A{pers + 1}"])):
                                        indind = pers
                                

                                
                                break
                            else: 
                                BlackList.append(index)

                        else: 
                            BlackList.append(index)
                            print(type(str(trans.translate_text(text=message.text, target_lang='RU')).upper()))
                            print(str(trans.translate_text(text=message.text, target_lang='RU')).upper())
                            print("=")
                            print(type(str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper()))
                            print(str(trans.translate_text(text=ws[f"D{index}"].value, target_lang='RU')).upper())
                    else: 
                        BlackList.append(index)
                        print("актив")
            
            if (Nothing): #нет анкет
                send_mes(message,mark=0, text=f"Ввозможно вы ввели неправильно, пожайлуста, введите професию еще раз")
                #for i in BlackList:
                #    print(f'{ws[f"B{i}"].value} {ws[f"C{i}"].value.upper()} == {prof.upper()} /// {ws[f"D{i}"].value.upper()} == {message.text.upper()} /// {str(ws[f"H{index}"].value)} == true')
                bot.register_next_step_handler(message, get_city_anket)
        else:
            bot.send_message(chat_id=message.chat.id, text="Простите, простите")








    if(message.from_user.username == Admin):
        if (text == "Добавить Анкету"):
            create_anket(message)
        if (text == "Увидеть все анкеты"):
            i = 0
            for anket in anket_list:
                bot.send_message(chat_id=message.chat.id, text=i)
                send_mes(message, mark=0, text=anket.user)
                #bot.send_message(chat_id=message.chat.id, text=anket.user)
                if (anket.active == True):
                    with open(anket.photo, "rb") as f:
                        with open(anket.video, "rb") as v:
                            media_group = []
                            media_group.append(telebot.types.InputMediaPhoto(f))
                            media_group.append(telebot.types.InputMediaVideo(v))
                            current_user = anket.user
                            bot.send_media_group(chat_id=message.chat.id, media = media_group)
                            send_mes(message, mark=0, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Активированая")
                            #bot.send_message(chat_id=message.chat.id, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Активированая")
                            i += 1
                if (anket.active == False):
                    with open(anket.photo, "rb") as f:
                        with open(anket.video, "rb") as v:
                            media_group = []
                            media_group.append(telebot.types.InputMediaPhoto(f))
                            media_group.append(telebot.types.InputMediaVideo(v))
                            current_user = anket.user
                            bot.send_media_group(chat_id=message.chat.id, media = media_group)
                            send_mes(message, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Деактивированая", mark=DoChange(message))
                            #bot.send_message(chat_id=message.chat.id, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Деактивированая", reply_markup= DoChange(user=message.from_user.username))
                            i += 1
        if(text == "Изменить анкету"):
            bot.send_message(chat_id=message.chat.id, text="Айди")
            bot.register_next_step_handler(message, Change_adm)
        if(text == "Загрузить анкеты из даты"):
            DataRead()
            bot.send_message(chat_id=message.chat.id, text="Дата загруженна успешна")
        if (text == "Увидеть все вакансии"):
            Adm_vac_seeall(message)
        if (text == "Удалить вакансию"):
            Adm_vac_delete(message)
        if(text == "Загрузить вакансии из даты"):
            DataReadVak()
            bot.send_message(chat_id=message.chat.id, text="Дата загруженна успешна")
    

def Change_langue(message):
    Change_Language(message)
    send_mes(message, mark=classic(message), text="Поздравляем! Вы поменяли язык")

def Change_adm (message):
    anket = anket_list[int(message.text)]
    bot.send_message(chat_id=message.chat.id, text=anket.user)
    if (anket.active == True):
        with open(anket.photo, "rb") as f:
            with open(anket.video, "rb") as v:
                media_group = []
                media_group.append(telebot.types.InputMediaPhoto(f))
                media_group.append(telebot.types.InputMediaVideo(v))
                current_user = anket.user
                bot.send_media_group(chat_id=message.chat.id, media = media_group)
                bot.send_message(chat_id=message.chat.id, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Активированая", reply_markup= DoChange(user=message.from_user.username))
    if (anket.active == False):
        with open(anket.photo, "rb") as f:
            with open(anket.video, "rb") as v:
                media_group = []
                media_group.append(telebot.types.InputMediaPhoto(f))
                media_group.append(telebot.types.InputMediaVideo(v))
                current_user = anket.user
                bot.send_media_group(chat_id=message.chat.id, media = media_group)
                bot.send_message(chat_id=message.chat.id, text=f"Имя: {anket.name}\n\nПрофесии: {anket.profession} \nГород: {anket.city} \n\nОписание: \n{anket.description} \n\n\nСтатус вашей анкеты: Деактивированая", reply_markup= DoChange(user=message.from_user.username))
    bot.send_message(chat_id=message.chat.id, text=f" Имя - 1 \nОписание - 2 \nФото - 3\nВидео - 4 \nПрофесия - 5\n Город - 6\n Статус - 7")
    bot.register_next_step_handler(message, Change_procc_adm, message.text)

    
def SeeAllVaks(message):
    idd = 0
    for vak in Vacantion_list:
        if (vak.user == message.from_user.username):
            idd += 1
            send_mes(message, mark=0, text=f"Вакансия ниже имеет номер {idd}")
            bot.send_message(chat_id=message.chat.id, reply_markup= AddVak(message),text=f"{vak.desc}\n\n{trans.translate_text(text= vak.desc, target_lang=Find_Language(message))}")
            


def Change_procc_adm(message, id):
    bot.send_message(chat_id=message.chat.id, text="На что изменить?")
    if (int(message.text) == 1):
        bot.register_next_step_handler(message, Procc_Change_name_adm, int(id))
    if (int(message.text) == 2):
        bot.register_next_step_handler(message, Procc_Change_desc_adm, int(id))
    if (int(message.text) == 3):
        bot.register_next_step_handler(message, Procc_Change_desc_adm, int(id))
    if (int(message.text) == 4):
        bot.register_next_step_handler(message, Procc_Change_photo_adm, int(id))
    if (int(message.text) == 5):
        bot.register_next_step_handler(message, Procc_Change_prof_adm, int(id))
    if (int(message.text) == 6):
        bot.register_next_step_handler(message, Procc_Change_city_adm, int(id))


def Procc_Change_name_adm(message, id):
    anket_list[id].name = message.text
    ws["B" + str(id + 2)] = anket_list[id].name
    wb.save(fn)
    bot.send_message(chat_id=message.chat.id, text="Изменинено")

def Procc_Change_desc_adm(message, id):
    anket_list[id].description = message.text
    ws["E" + str(id + 2)] = anket_list[id].description
    wb.save(fn)
    bot.send_message(chat_id=message.chat.id, text="Изменинено")

def Procc_Change_photo_adm(message, id):
    try:
        file_id = message.photo[-1].file_id
        file_info = bot.get_file(file_id)
        file = bot.download_file(file_info.file_path)
        with open(f"{file_id}.jpg", "wb") as f:
            f.write(file)
        anket_list[id].photo = file_id
        ws["F" + str(id + 2)] = anket_list[id].photo
        wb.save(fn)
        bot.send_message(chat_id=message.chat.id, text="Изменинено")
    except:
        bot.send_message(chat_id=message.chat.id, text="Пожайлуста, отправте ФОТО")
        bot.register_next_step_handler(message, Change_adm) 

def Procc_Change_video_adm(message, id):
    try:
        video_id = message.video.file_id
        video = bot.get_file(video_id)
        video_url = f"https://api.telegram.org/file/bot{bot.token}/{video.file_path}"
        response = requests.get(video_url)
        if response.status_code == 200:
            with open(f"{video_id}.mp4", "wb") as f:
                f.write(response.content)
        anket_list[id].video = video_id
        ws["G" + str(id + 2)] = anket_list[id].video
        wb.save(fn)
        bot.send_message(chat_id=message.chat.id, text="Изменинено")
    except:
        bot.send_message(chat_id=message.chat.id, text="Неправильный формат файла, отправтье фото ")
        bot.register_next_step_handler(message, Change_adm)


def Procc_Change_prof_adm(message, id):
    current = Profession_list.index(anket_list[id].profession)
    Profession_list[current] = message.text
    anket_list[id].profession = message.text
    ws["C" + str(id + 2)] = anket_list[id].profession
    wb.save(fn)
    bot.send_message(chat_id=message.chat.id, text="Изменинено") 

def Procc_Change_city_adm(message, id):
    current = City_list.index(anket_list[id].city)
    City_list[current] = message.text 
    anket_list[id].city = message.text
    ws["C" + str(id + 2)] = anket_list[id].profession
    wb.save(fn)
    bot.send_message(chat_id=message.chat.id, text="Изменинено") 


def remake_vac(message):
    send_mes(message, mark=0, text="Напишите, под каким номером находиться вакансия которую вы хотите удалить?")
    #bot.send_message(chat_id=message.chat.id, text="Напишите, под каким номером находиться вакансия которую вы хотите удалить?")
    bot.register_next_step_handler(message, remake_vac_procc)

def remake_vac_procc(message):
    current_vak = FindWhichVak(message, message.text)
    Vacantion_list.remove(current_vak)
    print(f"ну тип - {int(message.text)}")
    index = str(FindIndexVak(current_vak.user, int(message.text)) + 2)
    print(index)
    wsvak["E" + index] = "#deleted"
    wbvak.save(fnvak)
    send_mes(message, mark=classic(message), text="Поздравляем, вы удалили вакансию")
    #bot.send_message(chat_id=message.chat.id, text="Поздравляем, вы удалили вакансию", reply_markup=classic(message.from_user.username))

def change_procc_name(message):
    try:
        if (message.text != None):
            anket_now = IsHisAnket(message.from_user.username)
            anket_now.name = message.text
            ws["B" + str(anket_list.index(anket_now) + 2)] = anket_now.name
            wb.save(fn)
            send_mes(message, text="Поздравляем! Вы поменяли имя", mark=DoChange(message))
            #bot.send_message(chat_id=message.chat.id, text="Поздравляем! Вы поменяли имя", reply_markup =DoChange(message.from_user.username))
        else:
            send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
            bot.register_next_step_handler(message, change_procc_name)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, change_procc_name)

def change_procc_desc(message):
    try:
        if (message.text != None):
            anket_now = IsHisAnket(message.from_user.username)
            anket_now.description = message.text
            ws["E" + str(anket_list.index(anket_now) + 2)] = anket_now.description
            wb.save(fn)
            send_mes(message, text="Поздравляем! Вы поменяли описание", mark=DoChange(message))
            #bot.send_message(chat_id=message.chat.id, text="Поздравляем! Вы поменяли описание",reply_markup =DoChange(message.from_user.username))
        else:
            send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
            bot.register_next_step_handler(message, change_procc_desc)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, change_procc_desc)

def change_procc_photo(message):
    try:
        anket_now = IsHisAnket(message.from_user.username)
        file_id = message.photo[-1].file_id
        file_info = bot.get_file(file_id)
        file = bot.download_file(file_info.file_path)
        with open(f"{file_id}.jpg", "wb") as f:
            f.write(file)
        anket_now.photo = f"{file_id}.jpg"
        ws["F" + str(anket_list.index(anket_now) + 2)] = anket_now.photo
        wb.save(fn)
        send_mes(message, text="Поздравляем! Вы поменяли фото", reply_markup =DoChange(message))
        #bot.send_message(chat_id=message.chat.id, text="Поздравляем! Вы поменяли фото", reply_markup =DoChange(message.from_user.username))
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, change_procc_photo)


def change_procc_prof(message):
    try:
        if (message.text != None):
            anket_now = IsHisAnket(message.from_user.username)
            for i in range(len(Profession_list)):
                if (anket_now.profession == Profession_list[i]):
                    Profession_list[i] = message.text
            anket_now.profession = message.text
            ws["C" + str(anket_list.index(anket_now) + 2)] = anket_now.profession
            wb.save(fn)
            send_mes(message, text="Поздравляем! Вы поменяли професию", reply_markup =DoChange(message), mark=DoChange(message))
            #bot.send_message(chat_id=message.chat.id, text="Поздравляем! Вы поменяли професию", reply_markup =DoChange(message.from_user.username))
        else:
            send_mes(message, mark=0, text="Хмм, интересно, а теперь попробуйте написать текстом")
            #bot.send_message(chat_id=message.chat.id, text="Хмм, интересно, а теперь попробуйте написать текстом")
            bot.register_next_step_handler(message, change_procc_prof)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, change_procc_prof)

def change_procc_video(message):
    try:
        anket_now = IsHisAnket(message.from_user.username)
        video_id = message.video.file_id
        video = bot.get_file(video_id)
        video_url = f"https://api.telegram.org/file/bot{bot.token}/{video.file_path}"
        response = requests.get(video_url)
        if response.status_code == 200:
            with open(f"{video_id}.mp4", "wb") as f:
                f.write(response.content)
        anket_now.video = f"{video_id}.mp4"
        ws["G" + str(anket_list.index(anket_now) + 2)] = anket_now.video
        wb.save(fn)
        send_mes(message, text="Поздравляем! Вы поменяли видео", mark=DoChange(message))
        #bot.send_message(chat_id=message.chat.id, text="Поздравляем! Вы поменяли видео", reply_markup =DoChange(message.from_user.username))
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, change_procc_video)

def change_procc_city(message):
    try:
        if (message.text != None):
            anket_now = IsHisAnket(message.from_user.username)
            for i in range(len(City_list)):
                if (anket_now.city == City_list[i]):
                    City_list[i] = message.text
            anket_now.city = message.text
            ws["D" + str(anket_list.index(anket_now) + 2)] = anket_now.city
            wb.save(fn)
            send_mes(message, text="Поздравляем! Вы поменяли город", mark=DoChange(message))
            #bot.send_message(chat_id=message.chat.id, text="Поздравляем! Вы поменяли город", reply_markup =DoChange(message.from_user.username))
        else:
            send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
            bot.register_next_step_handler(message, change_procc_prof)
    except:
        send_mes(message, mark=classic(message), text="Неверно, попробуйте еще раз или чуть позже")
        bot.register_next_step_handler(message, change_procc_prof)

def Adm_vac_seeall(message):
    for index in range(int(ws["J1"].value)):
        print(index)
        bot.send_message(chat_id=message.chat.id, text=f'{index} \n{ws[f"A{str(index + 2)}"].value} \n{ws[f"B{str(index + 2)}"].value} \n{ws[f"C{str(index + 2)}"].value} \n{ws[f"D{str(index + 2)}"].value} \n{ws[f"E{str(index + 2)}"].value} \n{ws[f"H{str(index + 2)}"].value}')
        media_group = []
        with open(ws[f"F{index + 2}"].value, "rb") as f:
            media_group.append(telebot.types.InputMediaPhoto(f))
        with open(ws[f"G{index + 2}"].value, "rb") as v:
            media_group.append(telebot.types.InputMediaVideo(v))
        bot.send_media_group(chat_id=message.chat.id, media = media_group)



def Adm_vac_delete(message):
    send_mes(message, mark=0, text="Отправь айди вакансии")
    #bot.send_message(chat_id=message.chat.id, text="Отправь айди вакансии")
    bot.register_next_step_handler(message, Adm_vac_delete_procc)


def Adm_vac_delete_procc(message):
    indexx = int(message.text) 
    wsvak["E" + str(indexx + 2)] = "#deleted"
    wbvak.save(fnvak)    
    Vacantion_list.pop(indexx)
    bot.send_message(chat_id=message.chat.id, text="я кончил")


while True:
    try:
        bot.polling(non_stop=True, interval=0)
    except Exception as e:
        print(e)
        time.sleep(5)
        continue
